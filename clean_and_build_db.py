import os
import csv
import sqlite3
from openpyxl import load_workbook

# Paths
DATA_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "data"))
RAW_DATA_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
os.makedirs(DATA_DIR, exist_ok=True)
EXCEL_FILE = os.path.join(RAW_DATA_DIR, "0.Data Guidline (1).xlsx")
if not os.path.exists(EXCEL_FILE):
    EXCEL_FILE = os.path.join(RAW_DATA_DIR, "0.Data Guidline.xlsx")
DB_FILE = os.path.join(DATA_DIR, "gcontest.db")

# Sheet to CSV file mapping
MAPPING = {
    '1': ('Data_Customer', 'Data_Customer.csv'),
    '2': ('Data_Transaction', 'Data_Transaction.csv'),
    '3': ('Data_Activity', 'Data_Activity.csv'),
    '4': ('Data_Deposit', 'Data_Deposit.csv'),
    '5': ('Data_Lending', 'Data_Lending.csv'),
    '6': ('Data_Card', 'Data_Card.csv')
}

# Clean date strings without using heavy pandas/datetime parser
def clean_date(date_str):
    if not date_str:
        return None
    date_str = date_str.strip()
    if not date_str:
        return None
    # If format is already YYYY-MM-DD
    if '-' in date_str:
        return date_str.split()[0]
    # If format is M/D/YYYY H:MM or M/D/YYYY
    if '/' in date_str:
        parts = date_str.split()
        date_part = parts[0]
        subparts = date_part.split('/')
        if len(subparts) == 3:
            m, d, y = subparts
            try:
                return f"{int(y):04d}-{int(m):02d}-{int(d):02d}"
            except ValueError:
                return date_str
    return date_str

def clean_bool(val):
    if not val:
        return None
    val = val.strip().upper()
    if val == 'Y':
        return 1
    if val == 'N':
        return 0
    return None

def clean_int(val):
    if not val:
        return None
    try:
        return int(float(val))
    except ValueError:
        return None

def clean_float(val):
    if not val:
        return None
    try:
        return float(val)
    except ValueError:
        return None

CLEANERS = {
    # Dates
    'CLIENT_CREATE_DATE': clean_date,
    'DATE_OF_BIRTH': clean_date,
    'IB_REGISTER_DATE': clean_date,
    'TRANS_DATE': clean_date,
    'ACTIVITY_DATE': clean_date,
    'MONTH': clean_date,
    
    # Booleans
    'STAFF': clean_bool,
    'SMS': clean_bool,
    
    # Integers
    'CUSTOMER_NUMBER': clean_int,
    'TRANS_HOUR': clean_int,
    'TRANS_NO': clean_int,
    'ACTIVITY_HOUR': clean_int,
    'ACTIVITY_NO': clean_int,
    'COUNT_CA_ACCT': clean_int,
    'COUNT_TD_ACCT': clean_int,
    'COUNT_OF_LOAN': clean_int,
    'COUNT_CREDITCARD': clean_int,
    'COUNT_DEBITCARD': clean_int,
    
    # Floats
    'TRANS_AMOUNT': clean_float,
    'AVG_CA_BALANCE': clean_float,
    'AVG_TD_BALANCE': clean_float,
    'AVG_LOAN_AMOUNT': clean_float
}

def clean_legends():
    print("--- 1. Cleaning Legends ---")
    if not os.path.exists(EXCEL_FILE):
        print(f"Error: Excel file {EXCEL_FILE} not found.")
        return False
        
    wb = load_workbook(EXCEL_FILE, read_only=True)
    
    # Clean Master Data Dictionary (Sheet '0')
    sheet0 = wb['0']
    dict_csv = os.path.join(DATA_DIR, "legend_Data_Dictionary.csv")
    with open(dict_csv, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['Table Name', 'Description', 'Index'])
        for row_idx, row in enumerate(sheet0.iter_rows(values_only=True)):
            if row_idx < 4:  # Skip first 4 rows (including headers)
                continue
            if len(row) >= 5:
                table_name = row[1]
                desc = row[2]
                idx = row[4]
                if table_name and str(table_name).strip():
                    writer.writerow([
                        str(table_name).strip(),
                        str(desc).strip() if desc else '',
                        str(idx).strip() if idx else ''
                    ])
    print(f"Saved cleaned master dictionary to {dict_csv}")

    # Clean schemas (Sheets '1' to '6')
    for sheet_name, (table_name, _) in MAPPING.items():
        sheet = wb[sheet_name]
        legend_csv = os.path.join(DATA_DIR, f"legend_{table_name}.csv")
        with open(legend_csv, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Column', 'Column Description', 'Data Type'])
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_idx < 4:  # Skip first 4 rows (including headers)
                    continue
                if len(row) >= 4:
                    col = row[1]
                    col_desc = row[2]
                    dt = row[3]
                    if col and str(col).strip():
                        col_str = str(col).strip()
                        # Rename mismatching columns in Data_Transaction legend to match actual headers
                        if table_name == 'Data_Transaction':
                            if col_str == 'TRXN_LV1':
                                col_str = 'TRANS_LV1'
                            elif col_str == 'TRXN_LV2':
                                col_str = 'TRANS_LV2'
                        writer.writerow([
                            col_str,
                            str(col_desc).strip() if col_desc else '',
                            str(dt).strip() if dt else ''
                        ])
        print(f"Saved cleaned legend to {legend_csv}")
    return True

def process_and_load_table(table_name, csv_filename, conn):
    csv_path = os.path.join(RAW_DATA_DIR, csv_filename)
    if not os.path.exists(csv_path):
        print(f"Error: {csv_path} does not exist.")
        return

    print(f"Streaming and loading {table_name} from {csv_filename}...")
    
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        try:
            headers = next(reader)
        except StopIteration:
            print(f"Error: {csv_filename} is empty.")
            return
            
        headers = [h.strip() if h.strip() else f"col_{idx}" for idx, h in enumerate(headers)]
        
        # Build column types for schema definition
        col_types = []
        for h in headers:
            cleaner = CLEANERS.get(h)
            if cleaner == clean_date:
                t = "DATE"
            elif cleaner == clean_bool:
                t = "BOOLEAN"
            elif cleaner == clean_int:
                t = "INTEGER"
            elif cleaner == clean_float:
                t = "REAL"
            else:
                t = "TEXT"
            col_types.append(f'"{h}" {t}')
            
        cursor = conn.cursor()
        cursor.execute(f"DROP TABLE IF EXISTS {table_name}")
        cursor.execute(f"CREATE TABLE {table_name} ({', '.join(col_types)})")
        
        # Insert query template
        placeholders = ', '.join(['?'] * len(headers))
        quoted_headers = [f'"{h}"' for h in headers]
        insert_query = f"INSERT INTO {table_name} ({', '.join(quoted_headers)}) VALUES ({placeholders})"
        
        batch = []
        batch_size = 50000
        count = 0
        row_cleaners = [CLEANERS.get(h, lambda x: x.strip() if x else None) for h in headers]
        
        for row in reader:
            if not row:
                continue
            # Pad or slice row if lengths mismatch
            if len(row) < len(headers):
                row += [None] * (len(headers) - len(row))
            elif len(row) > len(headers):
                row = row[:len(headers)]
                
            cleaned_row = tuple(cleaner(val) for cleaner, val in zip(row_cleaners, row))
            batch.append(cleaned_row)
            
            if len(batch) >= batch_size:
                cursor.executemany(insert_query, batch)
                conn.commit()
                count += len(batch)
                batch = []
                print(f"  Loaded {count:,} rows...")
                
        if batch:
            cursor.executemany(insert_query, batch)
            conn.commit()
            count += len(batch)
            
        print(f"Finished loading {table_name}: {count:,} rows.")

def build_database():
    print("\n--- 2. Building SQLite Database ---")
    if os.path.exists(DB_FILE):
        print(f"Removing old database file {DB_FILE}...")
        os.remove(DB_FILE)
        
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # SQLite optimizations for fast writes
    cursor.execute("PRAGMA synchronous = OFF")
    cursor.execute("PRAGMA journal_mode = OFF")
    cursor.execute("PRAGMA cache_size = 10000")
    
    for _, (table_name, csv_filename) in MAPPING.items():
        process_and_load_table(table_name, csv_filename, conn)
        
    # Build indexes on CUSTOMER_NUMBER for high-performance joins
    print("\n--- 3. Creating Indexes ---")
    for table_name in MAPPING.values():
        t_name = table_name[0]
        cursor.execute(f"PRAGMA table_info({t_name})")
        columns = [col[1] for col in cursor.fetchall()]
        if 'CUSTOMER_NUMBER' in columns:
            print(f"Creating index on CUSTOMER_NUMBER for {t_name}...")
            cursor.execute(f"CREATE INDEX IF NOT EXISTS idx_{t_name}_cust ON {t_name} (CUSTOMER_NUMBER)")
            
    conn.commit()
    
    # Print summary stats
    print("\n--- Database Summary Statistics ---")
    for _, (t_name, _) in MAPPING.items():
        cursor.execute(f"SELECT COUNT(*) FROM {t_name}")
        count = cursor.fetchone()[0]
        print(f"Table '{t_name}': {count:,} rows")
        
    conn.close()
    print("\nDatabase construction complete.")

if __name__ == "__main__":
    if clean_legends():
        build_database()
